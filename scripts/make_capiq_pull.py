#!/usr/bin/env python
"""Build the Capital IQ pull specification workbook.

    make capiq        # or: python scripts/make_capiq_pull.py

Analyst consensus is the largest single data improvement available to this
study, and Capital IQ is the licence to hand. But the plug-in is driven from
Excel, one cell at a time, and the pull has a trap in it that is invisible in
the output: ``IQ_EPS_EST`` returns the consensus **as it stands today** unless
an ``asOfDate`` is supplied. Pull it without one and every quarter back to 2014
is stamped with a forecast formed after the actual was known.

So the defence has to live in the request, not the response. This script emits
a workbook with the as-of date already computed for every announcement -- the
last trading session strictly before the print -- so whoever runs the pull
never has to derive one, plus two live check formulas that flag a returned row
disagreeing with what was expected.

Sheets
------
``Read me first``       what to run, in what order, and what to check.
``1 Consensus pull``    one row per announcement, batched, with the as-of date.
``2 Delisted prices``   names the free price sources no longer carry.
``Mnemonics``           the function reference, so the sheet stands alone.

Batch 1 is a deliberately small pilot. The exact argument order of the ``CIQ``
worksheet function differs between plug-in versions, and discovering that on
several hundred rows is cheap where discovering it on eighteen thousand is not.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import numpy as np
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parents[1]
SRC = REPO_ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from earnings_engine.utils.calendar import default_calendar  # noqa: E402

#: Columns of sheet 1, in order. Their letters are referenced by the check
#: formulas, so reordering them means reworking `_formula_rows`.
PULL_COLUMNS = [
    "Batch",
    "Ticker",
    "Period End (expected)",
    "Fiscal Quarter",
    "Announced (ET)",
    "As Of Date",
    "Consensus EPS",
    "Estimate StdDev",
    "Num Estimates",
    "Returned Period End",
    "Returned Actual EPS",
    "Reported EPS (ours)",
    "Check: period",
    "Check: actual",
]

#: How far a returned period end may sit from the expected one before it is
#: worth a look. Fiscal quarters here are inferred from the announcement date,
#: so a few weeks of disagreement is ordinary and a whole quarter is not.
PERIOD_TOLERANCE_DAYS = 55

#: Half a cent. Below this a disagreement between the vendor's actual and ours
#: is rounding; above it, one of the two is describing a different figure.
EPS_TOLERANCE = 0.005

MNEMONICS = [
    (
        "IQ_EPS_EST",
        "Mean consensus diluted EPS estimate",
        "Returns the consensus as it stands TODAY unless asOfDate is passed. This is the trap.",
    ),
    (
        "IQ_EPS_EST_STDDEV",
        "Standard deviation of the individual estimates",
        "Denominator of analyst SUE. Zero or blank must yield NaN, not an infinite surprise.",
    ),
    (
        "IQ_EPS_NUM_EST",
        "Number of contributing estimates",
        "Thin coverage is a reason to distrust the dispersion, so it is worth carrying.",
    ),
    (
        "IQ_EPS_ACT",
        "Actual reported diluted EPS",
        "Pulled only so the pull can be checked against our own XBRL figure.",
    ),
    (
        "IQ_PERIODDATE",
        "Fiscal period end of the returned figure",
        "Checked against the period we expected; that check is column M.",
    ),
    (
        "IQ_CLOSEPRICE_ADJ",
        "Split and dividend adjusted close",
        "Sheet 2. Capital IQ retains delisted securities; the free sources do not.",
    ),
    ("IQ_OPENPRICE", "Daily open", "Sheet 2."),
    ("IQ_VOLUME", "Daily volume in shares", "Sheet 2."),
]


# --- the pieces, kept pure so they can be tested without a vendor terminal ---


def as_of_dates(announced_utc: pd.Series, exchange_tz: str = "America/New_York") -> pd.Series:
    """The last trading session strictly before each announcement.

    Uniform across BMO, AMC and intraday prints on purpose. An AMC release on
    session D lands after that session's close, so the consensus as of D is
    arguably still clean -- but "arguably" is not a standard worth adopting
    when the cost of the stricter rule is one session of staleness and the cost
    of the looser one is a forecast that has already seen the answer.
    """
    stamp = pd.to_datetime(announced_utc, utc=True, errors="coerce")
    local_day = stamp.dt.tz_convert(exchange_tz).dt.tz_localize(None).dt.normalize()
    cal = default_calendar()
    return pd.Series(
        [
            pd.NaT if pd.isna(day) else cal.previous_session(day, inclusive=False)
            for day in local_day
        ],
        index=announced_utc.index,
    )


def build_pull_rows(
    events: pd.DataFrame,
    reported_eps: pd.DataFrame | None = None,
    exchange_tz: str = "America/New_York",
) -> pd.DataFrame:
    """One row per announcement, with the as-of date already resolved."""
    needed = {"ticker", "announced_at_utc", "period_end"}
    absent = needed - set(events.columns)
    if absent:
        raise KeyError(f"events frame is missing {sorted(absent)}")

    keep = [c for c in events.columns if c in needed | {"fiscal_quarter"}]
    df = events.loc[:, keep].copy()
    df["ticker"] = df["ticker"].astype(str).str.upper()
    df["period_end"] = pd.to_datetime(df["period_end"], errors="coerce")
    announced = pd.to_datetime(df["announced_at_utc"], utc=True, errors="coerce")
    df["announced_et"] = announced.dt.tz_convert(exchange_tz).dt.tz_localize(None)
    df["as_of"] = as_of_dates(df["announced_at_utc"], exchange_tz)
    if "fiscal_quarter" not in df.columns:
        pe = df["period_end"]
        df["fiscal_quarter"] = (
            pe.dt.year.astype("Int64").astype(str)
            + "Q"
            + pe.dt.quarter.astype("Int64").astype(str)
        )

    df["reported_eps"] = np.nan
    if reported_eps is not None and not reported_eps.empty:
        r = reported_eps.copy()
        r["ticker"] = r["ticker"].astype(str).str.upper()
        r["period_end"] = pd.to_datetime(r["period_end"], errors="coerce")
        r = r.drop_duplicates(["ticker", "period_end"], keep="last")
        df = df.drop(columns=["reported_eps"]).merge(
            r[["ticker", "period_end", "eps"]].rename(columns={"eps": "reported_eps"}),
            on=["ticker", "period_end"],
            how="left",
        )

    # An announcement we cannot date is one we cannot ask for a consensus on.
    dropped = int((df["as_of"].isna() | df["period_end"].isna()).sum())
    if dropped:
        print(f"  skipped {dropped} event(s) with no usable announcement or period date")
    df = df.loc[df["as_of"].notna() & df["period_end"].notna()]
    return df.sort_values(["ticker", "period_end"]).reset_index(drop=True)


def assign_batches(
    rows: pd.DataFrame, pilot_tickers: int = 20, batch_size: int = 25
) -> pd.DataFrame:
    """Split by ticker, never mid-ticker, with batch 1 as the pilot.

    Batching by ticker rather than by row keeps a name's whole history in one
    export, so a ticker that comes back wrong is one file to redo rather than a
    seam running through several.
    """
    if pilot_tickers < 1 or batch_size < 1:
        raise ValueError("pilot_tickers and batch_size must both be at least 1")
    tickers = sorted(rows["ticker"].unique())
    batch_of: dict[str, int] = dict.fromkeys(tickers[:pilot_tickers], 1)
    for i, ticker in enumerate(tickers[pilot_tickers:]):
        batch_of[ticker] = 2 + i // batch_size
    out = rows.copy()
    out["batch"] = out["ticker"].map(batch_of).astype(int)
    return out.sort_values(["batch", "ticker", "period_end"]).reset_index(drop=True)


def delisted_names(universe: pd.DataFrame, prices: pd.DataFrame) -> pd.DataFrame:
    """Universe members the price store does not cover.

    This is the survivorship hole, itemised. A free price source carries what
    is listed today; a name that left the index by being acquired or delisted
    is exactly the name whose returns the study most needs and least has.
    Capital IQ retains them, which is the whole reason for this sheet.
    """
    have = {str(t).upper() for t in prices["ticker"].astype(str)}
    u = universe.copy()
    u["ticker"] = u["ticker"].astype(str).str.upper()
    u = u.drop_duplicates("ticker", keep="first")
    out = u.loc[~u["ticker"].isin(have)].copy()
    for col in ("start_date", "end_date"):
        out[col] = pd.to_datetime(out[col], errors="coerce")
    if "sector" not in out.columns:
        out["sector"] = pd.NA
    return (
        out[["ticker", "start_date", "end_date", "sector"]]
        .sort_values("ticker")
        .reset_index(drop=True)
    )


# --- workbook rendering -----------------------------------------------------


def _formula_rows(n_rows: int) -> list[dict[str, str]]:
    """The live cells, one dict per data row, keyed by column letter.

    Written as formulas rather than blanks so the sheet pulls and checks itself
    the moment it is opened with the plug-in loaded. The argument order of CIQ
    varies between plug-in versions -- that is precisely what the pilot batch
    is for, and the read-me says so.
    """
    out = []
    for i in range(n_rows):
        r = i + 2  # row 1 is the header
        out.append(
            {
                "G": f'=CIQ($B{r},"IQ_EPS_EST","IQ_CQ",$C{r},$F{r})',
                "H": f'=CIQ($B{r},"IQ_EPS_EST_STDDEV","IQ_CQ",$C{r},$F{r})',
                "I": f'=CIQ($B{r},"IQ_EPS_NUM_EST","IQ_CQ",$C{r},$F{r})',
                "J": f'=CIQ($B{r},"IQ_PERIODDATE","IQ_CQ",$C{r},$F{r})',
                "K": f'=CIQ($B{r},"IQ_EPS_ACT","IQ_CQ",$C{r})',
                "M": (
                    f'=IF(ISNUMBER($J{r}),IF(ABS($J{r}-$C{r})>{PERIOD_TOLERANCE_DAYS},'
                    f'"CHECK","ok"),"")'
                ),
                "N": (
                    f'=IF(AND(ISNUMBER($K{r}),ISNUMBER($L{r})),'
                    f'IF(ABS($K{r}-$L{r})>{EPS_TOLERANCE},"CHECK","ok"),"")'
                ),
            }
        )
    return out


README_LINES = [
    ("h", "Capital IQ pull specification"),
    (
        "p",
        "Generated by scripts/make_capiq_pull.py. Do not hand-edit: regenerate with `make capiq`.",
    ),
    ("b", ""),
    ("h", "Why the as-of date is the whole point"),
    (
        "p",
        "IQ_EPS_EST returns the consensus as it stands TODAY unless an asOfDate is supplied.",
    ),
    (
        "p",
        "Pull it without one and every historical quarter comes back stamped with a forecast",
    ),
    (
        "p",
        "formed after the actual was known -- a forecast that already contains its answer. The",
    ),
    (
        "p",
        "backtest that results is spectacular and worthless, and nothing in the output looks",
    ),
    ("p", "wrong. Column F has the date already computed. Pass it. Do not clear it."),
    ("b", ""),
    ("h", "Order of work"),
    (
        "p",
        "1. Run BATCH 1 only. It is a small pilot, and it exists because the argument order",
    ),
    ("p", "   of the CIQ worksheet function differs between plug-in versions. If the pilot comes"),
    ("p", "   back empty or shifted, fix the formula there -- not on eighteen thousand rows."),
    ("p", "2. Check columns M and N. Both should read 'ok'. M flags a returned period end more"),
    (
        "p",
        f"   than {PERIOD_TOLERANCE_DAYS} days from the one expected; N flags a returned actual",
    ),
    ("p", f"   disagreeing with our own XBRL figure by more than {EPS_TOLERANCE:.3f}."),
    ("p", "3. Only once the pilot is clean, run the remaining batches."),
    ("p", "4. Save each batch as CSV or XLSX into data/vendor/consensus/ (any number of files;"),
    ("p", "   they are concatenated). Nothing there is committed -- the data is licensed."),
    ("p", "5. Run: eee vendor-check --provider capitaliq"),
    ("p", "6. Re-run the study and compare sue_analyst against the existing time-series SUE."),
    ("b", ""),
    ("h", "Sheet 2: delisted prices"),
    ("p", "The universe is point-in-time, so it contains names that have since been acquired or"),
    ("p", "delisted. The free price sources no longer carry them, which biases the study toward"),
    ("p", "survivors. Capital IQ retains delisted securities, so this sheet is the substitute for"),
    ("p", "CRSP and it is how that bias gets closed rather than merely documented."),
    ("b", ""),
    ("h", "If a check fires in volume"),
    ("p", "Stop. A warning about snapshots post-dating the figures they forecast means the pull"),
    ("p", "was run without asOfDate. Fix the pull. Do not suppress the warning -- the pipeline"),
    ("p", "raises it because there is no way to repair the data after the fact."),
]


def write_workbook(
    path: Path,
    pull: pd.DataFrame,
    delisted: pd.DataFrame,
    price_start: str,
    price_end: str,
) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")

    wb = Workbook()

    # --- read me ---
    ws = wb.active
    ws.title = "Read me first"
    for i, (kind, text) in enumerate(README_LINES, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        if kind == "h":
            cell.font = bold
    ws.column_dimensions["A"].width = 100
    ws["A1"].font = Font(bold=True, size=14)

    # --- sheet 1: the pull ---
    ws = wb.create_sheet("1 Consensus pull")
    ws.append(PULL_COLUMNS)
    for cell in ws[1]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for _, row in pull.iterrows():
        ws.append(
            [
                int(row["batch"]),
                row["ticker"],
                row["period_end"].date(),
                row["fiscal_quarter"],
                row["announced_et"],
                row["as_of"].date(),
                None,
                None,
                None,
                None,
                None,
                None if pd.isna(row["reported_eps"]) else float(row["reported_eps"]),
                None,
                None,
            ]
        )
    for i, formulas in enumerate(_formula_rows(len(pull))):
        for letter, formula in formulas.items():
            ws[f"{letter}{i + 2}"] = formula
    widths = (7, 10, 20, 14, 20, 13, 14, 15, 14, 20, 18, 18, 13, 13)
    for letter, width in zip("ABCDEFGHIJKLMN", widths, strict=True):
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(PULL_COLUMNS))}{len(pull) + 1}"
    for letter in ("F",):
        ws[f"{letter}1"].fill = warn_fill

    # --- sheet 2: delisted prices ---
    ws = wb.create_sheet("2 Delisted prices")
    ws.append(["Ticker", "In index from", "In index to", "Sector", "Price history needed"])
    for cell in ws[1]:
        cell.font = bold
        cell.fill = header_fill
    for _, row in delisted.iterrows():
        ws.append(
            [
                row["ticker"],
                None if pd.isna(row["start_date"]) else row["start_date"].date(),
                None if pd.isna(row["end_date"]) else row["end_date"].date(),
                None if pd.isna(row["sector"]) else str(row["sector"]),
                f"{price_start} to {price_end}: IQ_OPENPRICE, IQ_CLOSEPRICE_ADJ, IQ_VOLUME (daily)",
            ]
        )
    for letter, width in zip("ABCDE", (10, 14, 14, 28, 68), strict=True):
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"

    # --- mnemonics ---
    ws = wb.create_sheet("Mnemonics")
    ws.append(["Mnemonic", "What it returns", "Note"])
    for cell in ws[1]:
        cell.font = bold
        cell.fill = header_fill
    for row in MNEMONICS:
        ws.append(list(row))
    for letter, width in zip("ABC", (24, 44, 92), strict=True):
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


# --- entry point ------------------------------------------------------------


def main(argv: list[str] | None = None) -> int:
    from earnings_engine.config import load_config
    from earnings_engine.data.base import ProviderError
    from earnings_engine.data.providers.local import LocalProvider

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", default="conf/config-real.yaml")
    parser.add_argument("--raw-dir", default=None, help="defaults to the config's data dir /raw")
    parser.add_argument("--out", default="docs/capiq-pull-specification.xlsx")
    parser.add_argument("--start", default="2014-06-01")
    parser.add_argument("--end", default=None, help="defaults to today")
    parser.add_argument("--pilot-tickers", type=int, default=20)
    parser.add_argument("--batch-size", type=int, default=25)
    args = parser.parse_args(argv)

    cfg = load_config(REPO_ROOT / args.config)
    raw_dir = Path(args.raw_dir) if args.raw_dir else cfg.resolved_paths().raw
    end = args.end or pd.Timestamp.today().strftime("%Y-%m-%d")
    provider = LocalProvider(raw_dir=raw_dir)

    try:
        universe = provider.get_universe()
        tickers = sorted(universe["ticker"].astype(str).str.upper().unique())
        events = provider.get_events(tickers, args.start, end)
        prices = provider.get_prices([*tickers, cfg.returns.market_symbol], args.start, end)
    except ProviderError as exc:
        print(f"cannot build the pull specification: {exc}", file=sys.stderr)
        print(
            "\nThis workbook is derived from the acquired data, so data/raw has to exist "
            "first:\n  eee download --start 2014-06-01",
            file=sys.stderr,
        )
        return 2

    reported = None
    try:
        fundamentals = provider.get_fundamentals(tickers, args.start, end)
        eps = fundamentals.loc[fundamentals["item"] == "eps_diluted"]
        reported = eps[["ticker", "period_end", "value"]].rename(columns={"value": "eps"})
    except ProviderError as exc:
        print(f"  no reported EPS to check against ({exc}); column L will be blank")

    rows = build_pull_rows(events, reported, exchange_tz=cfg.events.exchange_tz)
    pull = assign_batches(rows, args.pilot_tickers, args.batch_size)
    delisted = delisted_names(universe, prices)

    out = write_workbook(
        REPO_ROOT / args.out if not Path(args.out).is_absolute() else Path(args.out),
        pull,
        delisted,
        args.start,
        end,
    )
    n_batches = int(pull["batch"].max())
    pilot = pull.loc[pull["batch"] == 1]
    matched = int(pull["reported_eps"].notna().sum())
    print(f"wrote {out}")
    print(
        f"  sheet 1: {len(pull):,} rows x {len(PULL_COLUMNS)} cols, "
        f"{pull['ticker'].nunique()} tickers, {n_batches} batches"
    )
    print(f"  batch 1 pilot: {len(pilot):,} rows / {pilot['ticker'].nunique()} tickers")
    print(f"  reported EPS available for {matched:,}/{len(pull):,} rows (check column N)")
    print(f"  sheet 2: {len(delisted)} delisted name(s) with no price history")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
