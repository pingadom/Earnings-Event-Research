"""Tests for the Capital IQ pull specification.

The workbook is a request, not a result, so nothing here checks a number the
study depends on. What it checks is that the request cannot be issued wrong in
the one way that matters: an as-of date that is missing, or that falls on or
after the announcement it is supposed to precede.
"""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

import pandas as pd
import pytest

REPO_ROOT = Path(__file__).resolve().parents[1]


def _load_module():
    """Import the script by path -- scripts/ is not a package."""
    path = REPO_ROOT / "scripts" / "make_capiq_pull.py"
    spec = importlib.util.spec_from_file_location("make_capiq_pull", path)
    module = importlib.util.module_from_spec(spec)
    sys.modules["make_capiq_pull"] = module
    spec.loader.exec_module(module)
    return module


capiq = _load_module()


@pytest.fixture
def sample_events():
    """Four names, eight quarters each, announcing at a mix of hours."""
    rows = []
    for i, ticker in enumerate(["AAA", "BBB", "CCC", "DDD"]):
        for q in range(8):
            period_end = pd.Timestamp("2018-03-31") + pd.offsets.QuarterEnd(q)
            # A spread of BMO / intraday / AMC prints, ~5 weeks after period end.
            hour = (7, 12, 16, 20)[(i + q) % 4]
            announced = (period_end + pd.Timedelta(days=35)).replace(hour=hour)
            rows.append(
                {
                    "ticker": ticker,
                    "period_end": period_end,
                    "announced_at_utc": announced.tz_localize("America/New_York").tz_convert(
                        "UTC"
                    ),
                }
            )
    return pd.DataFrame(rows)


def test_the_as_of_date_always_precedes_the_announcement(sample_events):
    """The one property the whole workbook exists to guarantee."""
    rows = capiq.build_pull_rows(sample_events)
    assert len(rows) == len(sample_events)
    announced_day = rows["announced_et"].dt.normalize()
    assert (rows["as_of"] < announced_day).all()


def test_the_as_of_date_is_a_trading_session(sample_events):
    from earnings_engine.utils.calendar import default_calendar

    cal = default_calendar()
    rows = capiq.build_pull_rows(sample_events)
    assert all(cal.is_session(day) for day in rows["as_of"])


def test_a_monday_announcement_reaches_back_past_the_weekend():
    """A calendar-day offset would land on a Sunday and return nothing."""
    monday = pd.Timestamp("2024-04-15 07:00").tz_localize("America/New_York").tz_convert("UTC")
    events = pd.DataFrame(
        {"ticker": ["AAA"], "period_end": [pd.Timestamp("2024-03-31")], "announced_at_utc": [monday]}
    )
    rows = capiq.build_pull_rows(events)
    assert rows["as_of"].iloc[0] == pd.Timestamp("2024-04-12")  # the Friday


def test_an_undatable_event_is_dropped_not_guessed(sample_events):
    broken = sample_events.copy()
    broken.loc[broken.index[:3], "announced_at_utc"] = pd.NaT
    rows = capiq.build_pull_rows(broken)
    assert len(rows) == len(sample_events) - 3
    assert rows["as_of"].notna().all()


def test_batches_never_split_a_ticker(sample_events):
    rows = capiq.build_pull_rows(sample_events)
    batched = capiq.assign_batches(rows, pilot_tickers=1, batch_size=2)
    per_ticker = batched.groupby("ticker")["batch"].nunique()
    assert (per_ticker == 1).all()
    assert set(batched.loc[batched["batch"] == 1, "ticker"]) == {"AAA"}
    assert int(batched["batch"].max()) == 3  # 1 pilot + 3 names over batches of 2


def test_reported_eps_attaches_for_the_check_column(sample_events):
    reported = pd.DataFrame(
        {
            "ticker": ["AAA", "AAA"],
            "period_end": [pd.Timestamp("2018-03-31"), pd.Timestamp("2018-06-30")],
            "eps": [1.23, 1.44],
        }
    )
    rows = capiq.build_pull_rows(sample_events, reported)
    matched = rows.loc[rows["reported_eps"].notna()]
    assert len(matched) == 2
    assert set(matched["ticker"]) == {"AAA"}


def test_delisted_names_are_the_universe_minus_the_price_store():
    universe = pd.DataFrame(
        {
            "ticker": ["AAA", "BBB", "CCC"],
            "start_date": pd.to_datetime(["2014-01-01"] * 3),
            "end_date": pd.to_datetime(["2100-01-01", "2019-06-30", "2100-01-01"]),
            "sector": ["Tech", "Energy", "Health"],
        }
    )
    prices = pd.DataFrame({"ticker": ["AAA", "ccc", "SPY"]})
    out = capiq.delisted_names(universe, prices)
    assert list(out["ticker"]) == ["BBB"]


def test_the_workbook_wires_the_as_of_column_into_every_estimate_formula(sample_events, tmp_path):
    from openpyxl import load_workbook

    rows = capiq.assign_batches(capiq.build_pull_rows(sample_events), pilot_tickers=1, batch_size=2)
    delisted = pd.DataFrame(
        {
            "ticker": ["ZZZ"],
            "start_date": [pd.Timestamp("2014-01-01")],
            "end_date": [pd.Timestamp("2019-06-30")],
            "sector": ["Energy"],
        }
    )
    out = capiq.write_workbook(tmp_path / "spec.xlsx", rows, delisted, "2014-06-01", "2024-12-31")

    wb = load_workbook(out)
    assert wb.sheetnames == [
        "Read me first",
        "1 Consensus pull",
        "2 Delisted prices",
        "Mnemonics",
    ]
    ws = wb["1 Consensus pull"]
    assert [c.value for c in ws[1]] == capiq.PULL_COLUMNS
    assert ws.max_row == len(rows) + 1

    # Every consensus cell must reference column F. A formula that forgets it
    # silently returns today's consensus, which is the failure this guards.
    for r in range(2, ws.max_row + 1):
        for letter in ("G", "H", "I"):
            formula = ws[f"{letter}{r}"].value
            assert formula.startswith("=CIQ("), formula
            assert f"$F{r}" in formula, formula
        assert str(capiq.PERIOD_TOLERANCE_DAYS) in ws[f"M{r}"].value
        assert str(capiq.EPS_TOLERANCE) in ws[f"N{r}"].value
